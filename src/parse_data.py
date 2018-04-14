# -*- coding: utf-8 -*-
""" 
  将指定下的.xlsx文件 制作成标准的.line文件
"""
from openpyxl import load_workbook
from itertools import chain
import os
#将指定目录下所有文件（必须为.xlsx）写入一个文件.line文件
def parseToLine(dirPath):
	print(dirPath)
	for root, dirs, files in os.walk(dirPath):
		with open(os.path.join(dirPath,'parse.line'),'w+',encoding='utf-8') as lineFile:
			files = filter(lambda file:os.path.splitext(file)[1] == '.xlsx',files) 
			for file in files:
				filePath = os.path.join(root, file)
				print(filePath)
				workbook = load_workbook(filePath)
				sheets = workbook.sheetnames
				for sheet in sheets:
					booksheet = workbook[sheet]
					for val in chain.from_iterable(booksheet.values):
						print(val,file=lineFile)
#将有的三短信类作成.line文件
for froot, fdirs, ffiles in os.walk('data/'):
	for fdir in fdirs:
		parseToLine(os.path.join(froot, fdir))
